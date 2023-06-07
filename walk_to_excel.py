# file name: walk_to_excel.py
# writing time: 2023-06-07
# author: Fla5hback


import os
import openpyxl
from openpyxl.styles import Font


# 创建一个新的Excel工作簿
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "文件列表"

# 修改第一行字体颜色
worksheet = workbook.active
for i in range(1, 5):
    cell = worksheet.cell(row=1, column=i)
    cell.font = Font(size=15, color="0000CCFF", bold=True)


# 设置表头
sheet.cell(row=1, column=1, value="文件夹")
sheet.cell(row=1, column=2, value="文件名")
sheet.cell(row=1, column=3, value="文件类型")
sheet.cell(row=1, column=4, value="文件大小")

# 遍历目录下的所有文件
directory = "."  # 当前目录
row = 3
for foldername, subfolders, filenames in os.walk(directory):
    for filename in filenames:
        file_path = os.path.join(foldername,filename)
        file_size = os.path.getsize(file_path)
        file_type = os.path.splitext(filename)[-1][1:]

        # 将文件夹名、文件名和文件大小写入Excel
        sheet.cell(row=row, column=1, value=foldername)
        sheet.cell(row=row, column=2, value=filename)
        sheet.cell(row=row, column=3, value=file_type)
        sheet.cell(row=row, column=4, value=file_size)
        row += 1

# 保存Excel文件
workbook.save("文件列表.xlsx")